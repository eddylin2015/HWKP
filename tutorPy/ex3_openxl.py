"pip install openpyxl" 
import numpy as np
from IPython.display import Latex,HTML,Markdown
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
import datetime
wb=load_workbook(filename='table.xlsx')
wb.sheetnames
mcol=5; mrow=2827
ws=wb['工作表4']
data=[]
data_=[]
for idx,row in enumerate(ws.iter_rows(min_row=0,max_col=mcol,max_row=mrow)):
    r1=[]
    for i in range(0,mcol):
        v=row[i].value
        r1.append(v)
    data.append(r1)

print(data)